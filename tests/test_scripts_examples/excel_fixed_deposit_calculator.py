import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from utilities import xlutils

ops = webdriver.ChromeOptions()
ops.add_argument("--disable-notifications")

driver = webdriver.Chrome(options=ops)
time.sleep(5)
# driver.implicitly_wait(10)
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")

driver.maximize_window()
# save data
file = "/Users/vfurman/Desktop/caldata.xlsx"
rows = xlutils.getRowCount(file, "Sheet1")
# //*[@id="wzrk-confirm"]
# driver.find_element(By.XPATH, '//*[@id="wzrk-confirm"]').click()
time.sleep(5)

for r in range(2, rows+1):
    # reading data from excel
    pric = xlutils.readData(file, "Sheet1", r, 1)
    rateodinterest = xlutils.readData(file, "Sheet1", r, 2)
    period1 = xlutils.readData(file, "Sheet1", r, 3)
    period2 = xlutils.readData(file, "Sheet1", r, 4)
    fre = xlutils.readData(file, "Sheet1", r, 5)
    exp_mvalue = xlutils.readData(file, "Sheet1", r, 6)

    # passing data to application
    driver.find_element(By.XPATH, "//input[@id='principal']").send_keys(pric)
    driver.find_element(By.XPATH, "//input[@id='interest']").send_keys(rateodinterest)
    driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(period1)

    perioddrpd = Select(driver.find_element(By.XPATH, "//select[@id='tenurePeriod']"))
    perioddrpd.select_by_visible_text(period2)

    driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]/img").click()  # calculate button
    act_mvalue = driver.find_element(By.XPATH, "//span[@id='resp_matval']/strong").text

    # validation
    if float(exp_mvalue) == float(act_mvalue):
        print("test passed")
        xlutils.writeData(file, "Sheet1", r, 8, "Passed")
        xlutils.fillGreenColor(file, "Sheet1", r, 8)
    else:
        print("test failed")
        xlutils.writeData(file, "Sheet1", r, 8, "Failed")
        xlutils.fillRedColor(file, "Sheet1", r, 8)

    driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[2]/img").click()
    time.sleep(3)

driver.close()


# workbook = openpyxl.load_workbook(file)
# # sheet = workbook["Sheet1"]
# sheet = workbook.active  # get active sheet from excel
# # rows = sheet.max_row
# # cols = sheet.max_column
#
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "welcome"
#
# workbook.save(file)

# save multiple data

# file = "/Users/vfurman/Desktop/testdata.xlsx"
#
# workbook = openpyxl.load_workbook(file)
#
# sheet = workbook.active  # get active sheet from excel
#
# sheet.cell(1, 1).value = "123"
# sheet.cell(1, 2).value = "Smith"
# sheet.cell(1, 3).value = "Engineer"
#
# sheet.cell(2, 1).value = "567"
# sheet.cell(2, 2).value = "John"
# sheet.cell(2, 3).value = "Manager"
#
# sheet.cell(3, 1).value = "000"
# sheet.cell(3, 2).value = "David"
# sheet.cell(3, 3).value = "Developer"

# workbook.save(file)


