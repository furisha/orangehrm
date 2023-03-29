import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

file = "/Users/vfurman/Desktop/data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows = sheet.max_row
cols = sheet.max_column

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end='         ')
    print()


