import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

# save data with "welcome"
# file = "/Users/vfurman/Desktop/testdata.xlsx"
#
# workbook = openpyxl.load_workbook(file)
# # sheet = workbook["Sheet1"]
# sheet = workbook.active  # get active sheet from excel
# # rows = sheet.max_row
# # cols = sheet.max_column
#
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "welcome"
#
# workbook.save(file)

# save multiple data

file = "/Users/vfurman/Desktop/testdata.xlsx"

workbook = openpyxl.load_workbook(file)

sheet = workbook.active  # get active sheet from excel

sheet.cell(1, 1).value = "123"
sheet.cell(1, 2).value = "Smith"
sheet.cell(1, 3).value = "Engineer"

sheet.cell(2, 1).value = "567"
sheet.cell(2, 2).value = "John"
sheet.cell(2, 3).value = "Manager"

sheet.cell(3, 1).value = "000"
sheet.cell(3, 2).value = "David"
sheet.cell(3, 3).value = "Developer"

workbook.save(file)
print("All good")


